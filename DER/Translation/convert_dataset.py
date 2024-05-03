import openpyxl
import os
import shutil

def main(model, dataset, src, tgt):
    correct_idx = []
    count = 0
    root_report = "/home/changshu/CodeMind/Results/DER/Translation/translation_stats"
    path_report = os.path.join(root_report, f"{model}_{dataset}_compileReport_from_{src}_to_{tgt}_no_test.xlsx")
    dataframe = openpyxl.load_workbook(path_report)
    dataframe = dataframe.active
    for row in range(0, dataframe.max_row):
        col_idx = 0
        code_id, code_status = "", ""
        for col in dataframe.iter_cols(1, dataframe.max_column):
            if col_idx == 3:
                code_id = col[row].value
            if col_idx == 4:
                code_status = col[row].value
            col_idx += 1
        
        if code_status == "Correct":
            correct_idx.append(code_id)
        else:
            pass
    root_write = f"/home/changshu/CodeMind/dataset/Translation/{model}/{dataset}_{src}_{tgt}"
    root_read = f"/home/changshu/CodeMind/Results/DER/Translation/translation_output/no_test/{model}/{dataset}/{src}/{tgt}-sanitized"
    if dataset == "codenet" and src == "java":
        root_input_output = "/home/changshu/CodeMind/dataset/CodeNet/CodeNet-Java"
    elif dataset == "codenet" and src == "python":
        root_input_output = "/home/changshu/CodeMind/dataset/CodeNet/CodeNet-Python"
    elif dataset == "avatar" and src == "python":
        root_input_output = "/home/changshu/CodeMind/dataset/Avatar/Avatar-python"
    elif dataset == "avatar" and src == "java":
        root_input_output = "/home/changshu/CodeMind/dataset/Avatar/Avatar-java"   
    
    for d in correct_idx:
        if d == "Filename":
            continue
        code_index = d.split(".")[0]
        folder_wr = os.path.join(root_write,code_index)
        src_code = os.path.join(root_read, d)
        if not os.path.exists(src_code):
            print(src_code)
            continue
        if not os.path.exists(folder_wr):
            os.makedirs(folder_wr)
        
        if tgt == "python":
            tgt_code = os.path.join(folder_wr, "main.py")
        if tgt == "java":
            tgt_code = os.path.join(folder_wr, "Main.java")
        
        src_input= os.path.join(root_input_output, code_index, 'input.txt')
        tgt_input = os.path.join(folder_wr, 'input.txt')
        
        src_output = os.path.join(root_input_output, code_index, 'output.txt')
        tgt_output = os.path.join(folder_wr, 'output.txt')
        
        shutil.copy(src_code, tgt_code)
        shutil.copy(src_input, tgt_input)
        shutil.copy(src_output, tgt_output)
        

if __name__ == "__main__":
    model_list = ["codeqwen"]
    for m in model_list:
        # main(m, 'avatar', 'java', 'python')
        # main(m, 'avatar', 'python', 'java')
        main(m, 'codenet', 'java', 'python')
        main(m, 'codenet', 'python', 'java')