import unittest
import openpyxl


class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        data = []
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            workbook.close()
            return data
        except:
            return None

    def write_excel(self, data, file_name):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            workbook.close()
            return 1
        except:
            return 0

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if data is None or N >= len(data[0]):
            return 0
        new_data = []
        for row in data:
            new_row = list(row[:])
            if not str(row[N]).isdigit():
                new_row.append(str(row[N]).upper())
            else:
                new_row.append(row[N])
            new_data.append(new_row)
        new_file_name = save_file_name.split('.')[0] + '_process.xlsx'
        success = self.write_excel(new_data, new_file_name)
        return success, new_file_name
class Test(unittest.TestCase):
    def test(self):
            processor = ExcelProcessor()
            new_data = [
                ('Name', 'Age', 'Country'),
                ('John', 25, 'USA')
            ]
            save_file_name = ''
            success = processor.write_excel(new_data, save_file_name)
            return success